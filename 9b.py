import openpyxl
# Read data from a spreadsheet
wo = openpyxl.load_workbook('data.xlsx')
wos = wo.active
# Read values from cells
value1 = wos['A1'].value
value2 = wos.cell(row=2, column=1).value
# Display the read values
print("Read values from the spreadsheet:")
print("Value 1:", value1)
print("Value 2:", value2)
# Write data to the spreadsheet
wos['B1'] = "Hello"
wos.cell(row=2, column=2, value="World")
# Save the modified spreadsheet
wo.save('data.xlsx')
print("Data written to the spreadsheet successfully!")


